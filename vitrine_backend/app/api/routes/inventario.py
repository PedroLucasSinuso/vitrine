"""
app/api/routes/inventario.py  — substitui o arquivo existente na íntegra

Mudanças em relação ao original:
  • GET /admin/inventario/sessoes/{id}/exportar-excel  (NOVO)
    Exporta Excel com duas abas:
      - "Contagem"          → dados bipados da sessão (consolidado por código)
      - "Delta (vs. Sistema)" → cruza com estoque atual do SQLite (campo stock de Product)
    Requer role supervisor ou admin.
  • GET /admin/inventario/consolidado-geral/exportar-excel  (NOVO)
    Mesmo formato, mas consolida TODAS as sessões ativas.
  • Todo o resto permanece idêntico ao original.
"""

import io
import secrets
import logging
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, delete
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.api.deps import get_db, get_current_user, require_supervisor
from app.schemas.inventario_schema import (
    CriarSessaoInput,
    EntrarSessaoInput,
    SessaoResponse,
    ItemInventarioSubmit,
    ItemInventarioResponse,
    AtualizarItemInput,
)
from app.domain.models.inventario import SessaoInventario, ItemInventario
from app.domain.models.usuario import Usuario
from app.infrastructure.db.bootstrap import init_db

# Produto traz o estoque atual do SQLite
from app.domain.models.produto import Produto

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/admin/inventario", tags=["Inventario"])


# ─────────────────────────────────────────
# Helpers internos
# ─────────────────────────────────────────

def gerar_codigo_convite() -> str:
    return secrets.token_hex(3).upper()


def get_session_or_404(sessao_id: int, db: Session) -> SessaoInventario:
    sessao = db.execute(
        select(SessaoInventario).where(SessaoInventario.id == sessao_id)
    ).scalar_one_or_none()
    if not sessao:
        raise HTTPException(status_code=404, detail="Sessão não encontrada")
    return sessao


def require_sessao_ativa(sessao: SessaoInventario) -> None:
    """Levanta 400 se a sessão não estiver ativa.
    
    Chamado pelas rotas que modificam itens (adicionar, listar, atualizar, limpar)
    para impedir operações em sessões encerradas. A rota entrar_sessao faz esta
    verificação diretamente; as demais usam este helper."""
    if sessao.status != "ativa":
        raise HTTPException(status_code=400, detail="Sessão já encerrada")


def _build_sessao_response(sessao: SessaoInventario, db: Session) -> SessaoResponse:
    total_operadores = db.execute(
        select(func.count(func.distinct(ItemInventario.usuario_id)))
        .where(ItemInventario.sessao_id == sessao.id)
    ).scalar() or 0

    total_itens = db.execute(
        select(func.count(ItemInventario.id))
        .where(ItemInventario.sessao_id == sessao.id)
    ).scalar() or 0

    criador = db.execute(
        select(Usuario).where(Usuario.id == sessao.criado_por_id)
    ).scalar_one_or_none()

    return SessaoResponse(
        id=sessao.id,
        nome=sessao.nome,
        status=sessao.status,
        codigo_convite=sessao.codigo_convite,
        criado_por=criador.username if criador else "?",
        criado_em=sessao.criado_em,
        total_operadores=total_operadores,
        total_itens=total_itens,
    )


def _consolidar_itens_sessao(sessao_id: int, db: Session) -> list[dict]:
    """Retorna itens consolidados (sum por código) de UMA sessão."""
    rows = db.execute(
        select(
            ItemInventario.codigo,
            ItemInventario.nome,
            ItemInventario.grupo,
            ItemInventario.familia,
            func.sum(ItemInventario.quantidade).label("quantidade"),
        )
        .where(ItemInventario.sessao_id == sessao_id)
        .group_by(
            ItemInventario.codigo,
            ItemInventario.nome,
            ItemInventario.grupo,
            ItemInventario.familia,
        )
    ).all()
    return [
        {"codigo": r.codigo, "nome": r.nome, "grupo": r.grupo,
         "familia": r.familia, "quantidade": int(r.quantidade)}
        for r in rows
    ]


def _consolidar_itens_todas_sessoes(db: Session) -> list[dict]:
    """Retorna itens consolidados de TODAS as sessões ativas."""
    rows = db.execute(
        select(
            ItemInventario.codigo,
            ItemInventario.nome,
            ItemInventario.grupo,
            ItemInventario.familia,
            func.sum(ItemInventario.quantidade).label("quantidade"),
        )
        .join(SessaoInventario, ItemInventario.sessao_id == SessaoInventario.id)
        .where(SessaoInventario.status == "ativa")
        .group_by(
            ItemInventario.codigo,
            ItemInventario.nome,
            ItemInventario.grupo,
            ItemInventario.familia,
        )
    ).all()
    return [
        {"codigo": r.codigo, "nome": r.nome, "grupo": r.grupo,
         "familia": r.familia, "quantidade": int(r.quantidade)}
        for r in rows
    ]


# ─────────────────────────────────────────
# Gerador de Excel
# ─────────────────────────────────────────

_HEADER_FILL  = PatternFill("solid", start_color="1E3A5F")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_ALT_FILL     = PatternFill("solid", start_color="F4F7FB")
_TOTAL_FILL   = PatternFill("solid", start_color="E8F0FE")
_THIN         = Side(style="thin", color="D0D0D0")
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_RED_FONT     = Font(name="Arial", size=10, color="C0392B", bold=True)
_GREEN_FONT   = Font(name="Arial", size=10, color="1A7A3C", bold=True)
_NORMAL_FONT  = Font(name="Arial", size=10)


def _header_cell(ws, row: int, col: int, value: str):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _HEADER_FONT
    c.fill = _HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _BORDER
    return c


def _data_cell(ws, row: int, col: int, value, center: bool = False, alt: bool = False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _NORMAL_FONT
    c.border = _BORDER
    c.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
    )
    if alt:
        c.fill = _ALT_FILL
    return c


def _build_excel(
    itens_contados: list[dict],
    estoque_db: dict[str, float],  # codigo → estoque atual
    nome_relatorio: str,
    observacoes: list[dict] | None = None,
) -> bytes:
    """
    Gera o .xlsx em memória e retorna os bytes.

    itens_contados : lista de {codigo, nome, grupo, familia, quantidade}
    estoque_db     : mapa codigo → estoque atual do sistema
    observacoes    : lista opcional de {codigo, nome, observacao}
                     (se vazia, a aba "Observações" não é criada)
    """
    wb = Workbook()

    # ── Aba 1: Contagem ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Contagem"
    ws1.freeze_panes = "A2"

    cols1 = ["Código", "Produto", "Grupo", "Família", "Qtd. Contada"]
    for col, h in enumerate(cols1, 1):
        _header_cell(ws1, 1, col, h)
    ws1.row_dimensions[1].height = 22

    for ri, item in enumerate(itens_contados, 2):
        alt = ri % 2 == 0
        _data_cell(ws1, ri, 1, item["codigo"],     center=True,  alt=alt)
        _data_cell(ws1, ri, 2, item["nome"],                     alt=alt)
        _data_cell(ws1, ri, 3, item["grupo"],                    alt=alt)
        _data_cell(ws1, ri, 4, item["familia"],                  alt=alt)
        _data_cell(ws1, ri, 5, item["quantidade"], center=True,  alt=alt)

    last1 = len(itens_contados) + 1
    total_row1 = last1 + 1
    for col in range(1, 6):
        c = ws1.cell(row=total_row1, column=col)
        c.border = _BORDER
        c.fill = _TOTAL_FILL
        c.font = Font(bold=True, name="Arial", size=10)
        if col == 4:
            c.value = "TOTAL"
            c.alignment = Alignment(horizontal="right", vertical="center")
        elif col == 5:
            c.value = f"=SUM(E2:E{last1})"
            c.alignment = Alignment(horizontal="center", vertical="center")

    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 34
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 16
    ws1.column_dimensions["E"].width = 14

    # ── Aba 2: Delta ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Delta (vs. Sistema)")
    ws2.freeze_panes = "A2"

    cols2 = ["Código", "Produto", "Grupo", "Família",
             "Estoque Sistema", "Qtd. Contada", "Delta", "Status"]
    for col, h in enumerate(cols2, 1):
        _header_cell(ws2, 1, col, h)
    ws2.row_dimensions[1].height = 22

    for ri, item in enumerate(itens_contados, 2):
        alt = ri % 2 == 0
        cod = item["codigo"]
        contado = item["quantidade"]
        db_stock = estoque_db.get(cod)  # None se não encontrado no DB

        db_val: float | str = round(db_stock, 3) if db_stock is not None else "—"
        delta_formula = f"=F{ri}-E{ri}" if db_stock is not None else "—"

        if db_stock is None:
            status = "Sem cadastro"
        elif contado == 0:
            status = "Não contado"
        elif contado > db_stock:
            status = "Sobra"
        elif contado < db_stock:
            status = "Falta"
        else:
            status = "OK"

        _data_cell(ws2, ri, 1, cod,         center=True, alt=alt)
        _data_cell(ws2, ri, 2, item["nome"],              alt=alt)
        _data_cell(ws2, ri, 3, item["grupo"],             alt=alt)
        _data_cell(ws2, ri, 4, item["familia"],           alt=alt)
        _data_cell(ws2, ri, 5, db_val,      center=True, alt=alt)
        _data_cell(ws2, ri, 6, contado,     center=True, alt=alt)

        # Célula delta — fórmula colorida
        c_delta = ws2.cell(row=ri, column=7, value=delta_formula)
        c_delta.border = _BORDER
        c_delta.alignment = Alignment(horizontal="center", vertical="center")
        if alt:
            c_delta.fill = _ALT_FILL
        if db_stock is not None:
            delta = contado - db_stock
            if delta > 0:
                c_delta.font = _GREEN_FONT
            elif delta < 0:
                c_delta.font = _RED_FONT
            else:
                c_delta.font = Font(name="Arial", size=10, color="555555")
        else:
            c_delta.font = Font(name="Arial", size=10, color="AAAAAA")

        # Célula status — colorida
        c_status = ws2.cell(row=ri, column=8, value=status)
        c_status.border = _BORDER
        c_status.alignment = Alignment(horizontal="center", vertical="center")
        if alt:
            c_status.fill = _ALT_FILL
        if status == "OK":
            c_status.font = _GREEN_FONT
        elif status in ("Falta", "Não contado", "Sem cadastro"):
            c_status.font = _RED_FONT
        elif status == "Sobra":
            c_status.font = Font(name="Arial", size=10, color="D4800A", bold=True)
        else:
            c_status.font = _NORMAL_FONT

    # Totais da aba Delta
    last2 = len(itens_contados) + 1
    total_row2 = last2 + 1
    numeric_rows = [
        i for i, item in enumerate(itens_contados, 2)
        if estoque_db.get(item["codigo"]) is not None
    ]
    for col in range(1, 9):
        c = ws2.cell(row=total_row2, column=col)
        c.border = _BORDER
        c.fill = _TOTAL_FILL
        c.font = Font(bold=True, name="Arial", size=10)
        if col == 4:
            c.value = "TOTAL"
            c.alignment = Alignment(horizontal="right", vertical="center")
        elif col == 5:
            c.value = f"=SUM(E2:E{last2})"
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 6:
            c.value = f"=SUM(F2:F{last2})"
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 7:
            c.value = f"=SUM(G2:G{last2})"
            c.alignment = Alignment(horizontal="center", vertical="center")

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 34
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 17
    ws2.column_dimensions["F"].width = 15
    ws2.column_dimensions["G"].width = 10
    ws2.column_dimensions["H"].width = 15

    # ── Aba 3: Observações ──────────────────────────────────────────────
    if observacoes:
        ws3 = wb.create_sheet("Observações")
        ws3.freeze_panes = "A2"

        cols3 = ["Código", "Produto", "Observação"]
        for col, h in enumerate(cols3, 1):
            _header_cell(ws3, 1, col, h)
        ws3.row_dimensions[1].height = 22

        for ri, obs in enumerate(observacoes, 2):
            alt = ri % 2 == 0
            _data_cell(ws3, ri, 1, obs.get("codigo", ""), center=True, alt=alt)
            _data_cell(ws3, ri, 2, obs.get("nome", ""), alt=alt)
            _data_cell(ws3, ri, 3, obs.get("observacao", ""), alt=alt)

        ws3.column_dimensions["A"].width = 18
        ws3.column_dimensions["B"].width = 34
        ws3.column_dimensions["C"].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _get_estoque_db(
    codigos: list[str],
    db: Session,
) -> dict[str, float]:
    """
    Busca o estoque atual dos produtos no SQLite.
    Usa WHERE IN para buscar apenas os códigos necessários.
    Retorna mapa codigo_chamada → estoque.
    Produtos não encontrados ficam ausentes do mapa.
    """
    if not codigos:
        return {}
    try:
        rows = db.execute(
            select(Produto.codigo_chamada, Produto.estoque)
            .where(Produto.codigo_chamada.in_(codigos))
        ).all()
        return {r.codigo_chamada: float(r.estoque) for r in rows}
    except Exception:
        logger.exception("Falha ao buscar estoque para delta — exportando sem delta")
        return {}


# ─────────────────────────────────────────
# Endpoints existentes (sem alteração)
# ─────────────────────────────────────────

@router.get("/sessoes", response_model=list[SessaoResponse])
def listar_sessoes(
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(get_current_user),
):
    init_db()
    stmt = select(SessaoInventario).where(SessaoInventario.status == "ativa")
    sessoes = db.execute(stmt).scalars().all()
    return [_build_sessao_response(s, db) for s in sessoes]


@router.post("/sessoes", response_model=SessaoResponse, status_code=201)
def criar_sessao(
    body: CriarSessaoInput,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(require_supervisor),
):
    init_db()
    codigo = gerar_codigo_convite()
    while db.execute(
        select(SessaoInventario).where(SessaoInventario.codigo_convite == codigo)
    ).scalar_one_or_none():
        codigo = gerar_codigo_convite()

    sessao = SessaoInventario(
        nome=body.nome,
        criado_por_id=usuario.id,
        status="ativa",
        codigo_convite=codigo,
        criado_em=datetime.now(timezone.utc),
    )
    db.add(sessao)
    db.commit()
    db.refresh(sessao)
    return _build_sessao_response(sessao, db)


@router.post("/sessoes/entrar", response_model=SessaoResponse, status_code=201)
def entrar_sessao(
    body: EntrarSessaoInput,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(get_current_user),
):
    init_db()
    sessao = db.execute(
        select(SessaoInventario).where(SessaoInventario.codigo_convite == body.codigo_convite)
    ).scalar_one_or_none()
    if not sessao:
        raise HTTPException(status_code=404, detail="Sessão não encontrada")
    if sessao.status != "ativa":
        raise HTTPException(status_code=400, detail="Sessão já encerrada")
    return _build_sessao_response(sessao, db)


@router.patch("/sessoes/{sessao_id}", response_model=SessaoResponse)
def encerrar_sessao(
    sessao_id: int,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(require_supervisor),
):
    init_db()
    sessao = get_session_or_404(sessao_id, db)
    if sessao.criado_por_id != usuario.id:
        raise HTTPException(status_code=403, detail="Apenas o criador pode encerrar a sessão")
    sessao.status = "encerrada"
    sessao.encerrado_em = datetime.now(timezone.utc)
    db.commit()
    db.refresh(sessao)
    return _build_sessao_response(sessao, db)


@router.get("/sessoes/{sessao_id}/itens", response_model=list[ItemInventarioResponse])
def listar_itens(
    sessao_id: int,
    consolidado: bool = False,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(get_current_user),
):
    init_db()
    sessao = get_session_or_404(sessao_id, db)
    require_sessao_ativa(sessao)

    if consolidado and usuario.role in ("supervisor", "admin"):
        rows = db.execute(
            select(
                ItemInventario.codigo,
                ItemInventario.nome,
                ItemInventario.grupo,
                ItemInventario.familia,
                func.sum(ItemInventario.quantidade).label("quantidade"),
            )
            .where(ItemInventario.sessao_id == sessao_id)
            .group_by(
                ItemInventario.codigo,
                ItemInventario.nome,
                ItemInventario.grupo,
                ItemInventario.familia,
            )
        ).all()
        return [
            ItemInventarioResponse(
                codigo=r.codigo, nome=r.nome, grupo=r.grupo,
                familia=r.familia, quantidade=r.quantidade,
            )
            for r in rows
        ]

    rows = db.execute(
        select(ItemInventario)
        .where(ItemInventario.sessao_id == sessao_id)
        .where(ItemInventario.usuario_id == usuario.id)
    ).scalars().all()
    return [
        ItemInventarioResponse(
            codigo=r.codigo, nome=r.nome, grupo=r.grupo,
            familia=r.familia, quantidade=r.quantidade,
            observacao=r.observacao or "",
        )
        for r in rows
    ]


@router.post("/sessoes/{sessao_id}/itens", status_code=201)
def adicionar_item(
    sessao_id: int,
    body: ItemInventarioSubmit,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(get_current_user),
):
    init_db()
    sessao = get_session_or_404(sessao_id, db)
    require_sessao_ativa(sessao)

    try:
        existing = db.execute(
            select(ItemInventario)
            .where(ItemInventario.sessao_id == sessao_id)
            .where(ItemInventario.usuario_id == usuario.id)
            .where(ItemInventario.codigo == body.codigo)
        ).scalar_one_or_none()

        if existing:
            existing.quantidade += body.quantidade
            if body.observacao:
                existing.observacao = (
                    (existing.observacao + " | " + body.observacao)
                    if existing.observacao
                    else body.observacao
                )
        else:
            item = ItemInventario(
                sessao_id=sessao_id,
                usuario_id=usuario.id,
                codigo=body.codigo,
                nome=body.nome,
                grupo=body.grupo,
                familia=body.familia,
                quantidade=body.quantidade,
                observacao=body.observacao or None,
            )
            db.add(item)

        db.commit()
    except IntegrityError:
        db.rollback()
        existing = db.execute(
            select(ItemInventario)
            .where(ItemInventario.sessao_id == sessao_id)
            .where(ItemInventario.usuario_id == usuario.id)
            .where(ItemInventario.codigo == body.codigo)
        ).scalar_one_or_none()
        if existing:
            existing.quantidade += body.quantidade
            if body.observacao:
                existing.observacao = (
                    (existing.observacao + " | " + body.observacao)
                    if existing.observacao
                    else body.observacao
                )
        else:
            item = ItemInventario(
                sessao_id=sessao_id,
                usuario_id=usuario.id,
                codigo=body.codigo,
                nome=body.nome,
                grupo=body.grupo,
                familia=body.familia,
                quantidade=body.quantidade,
                observacao=body.observacao or None,
            )
            db.add(item)
        db.commit()

    return {"ok": True}


@router.get("/consolidado-geral", response_model=list[ItemInventarioResponse])
def consolidado_geral(
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(require_supervisor),
):
    init_db()
    rows = db.execute(
        select(
            ItemInventario.codigo,
            ItemInventario.nome,
            ItemInventario.grupo,
            ItemInventario.familia,
            func.sum(ItemInventario.quantidade).label("quantidade"),
        )
        .join(SessaoInventario, ItemInventario.sessao_id == SessaoInventario.id)
        .where(SessaoInventario.status == "ativa")
        .group_by(
            ItemInventario.codigo,
            ItemInventario.nome,
            ItemInventario.grupo,
            ItemInventario.familia,
        )
    ).all()
    return [
        ItemInventarioResponse(
            codigo=r.codigo, nome=r.nome, grupo=r.grupo,
            familia=r.familia, quantidade=r.quantidade,
        )
        for r in rows
    ]


@router.patch("/sessoes/{sessao_id}/itens/{codigo}")
def atualizar_item(
    sessao_id: int,
    codigo: str,
    body: AtualizarItemInput,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(get_current_user),
):
    init_db()
    sessao = get_session_or_404(sessao_id, db)
    require_sessao_ativa(sessao)

    item = db.execute(
        select(ItemInventario)
        .where(ItemInventario.sessao_id == sessao_id)
        .where(ItemInventario.usuario_id == usuario.id)
        .where(ItemInventario.codigo == codigo)
    ).scalar_one_or_none()

    if not item:
        raise HTTPException(status_code=404, detail="Item não encontrado")

    if body.quantidade <= 0:
        db.delete(item)
    else:
        item.quantidade = body.quantidade
        if body.observacao is not None:
            item.observacao = body.observacao or None

    db.commit()
    return {"ok": True}


@router.delete("/sessoes/{sessao_id}/itens")
def limpar_itens(
    sessao_id: int,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(get_current_user),
):
    init_db()
    sessao = get_session_or_404(sessao_id, db)
    require_sessao_ativa(sessao)

    db.execute(
        delete(ItemInventario)
        .where(ItemInventario.sessao_id == sessao_id)
        .where(ItemInventario.usuario_id == usuario.id)
    )
    db.commit()
    return {"ok": True}


# ─────────────────────────────────────────
# NOVOS endpoints de exportação Excel
# ─────────────────────────────────────────

@router.get("/sessoes/{sessao_id}/exportar-excel")
def exportar_excel_sessao(
    sessao_id: int,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(require_supervisor),
):
    """
    Exporta Excel de UMA sessão (consolidado de todos os operadores).
    Aba "Contagem"           → itens bipados, somados por código.
    Aba "Delta (vs. Sistema)" → cruza com estoque atual do SQLite.
    Apenas supervisores e admins.
    """
    init_db()
    sessao = get_session_or_404(sessao_id, db)

    itens = _consolidar_itens_sessao(sessao_id, db)
    if not itens:
        raise HTTPException(status_code=404, detail="Nenhum item nesta sessão")

    codigos = [i["codigo"] for i in itens]
    estoque_db = _get_estoque_db(codigos, db)

    # Observações desta sessão
    obs_rows = db.execute(
        select(ItemInventario.codigo, ItemInventario.nome, ItemInventario.observacao)
        .where(ItemInventario.sessao_id == sessao_id)
        .where(ItemInventario.observacao.isnot(None))
        .where(ItemInventario.observacao != "")
    ).all()
    observacoes = [{"codigo": r.codigo, "nome": r.nome, "observacao": r.observacao} for r in obs_rows]

    nome = sessao.nome or f"sessao_{sessao_id}"
    nome_arquivo = f"inventario_{nome}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    nome_arquivo = "".join(c for c in nome_arquivo if c.isalnum() or c in "-_.")

    excel_bytes = _build_excel(itens, estoque_db, nome, observacoes)

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


@router.get("/consolidado-geral/exportar-excel")
def exportar_excel_consolidado_geral(
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(require_supervisor),
):
    """
    Exporta Excel consolidado de TODAS as sessões ativas.
    Mesmo formato de duas abas.
    """
    init_db()

    itens = _consolidar_itens_todas_sessoes(db)
    if not itens:
        raise HTTPException(status_code=404, detail="Nenhum item nas sessões ativas")

    codigos = [i["codigo"] for i in itens]
    estoque_db = _get_estoque_db(codigos, db)

    # Observações de todas as sessões ativas
    obs_rows = db.execute(
        select(ItemInventario.codigo, ItemInventario.nome, ItemInventario.observacao)
        .join(SessaoInventario, ItemInventario.sessao_id == SessaoInventario.id)
        .where(SessaoInventario.status == "ativa")
        .where(ItemInventario.observacao.isnot(None))
        .where(ItemInventario.observacao != "")
    ).all()
    observacoes = [{"codigo": r.codigo, "nome": r.nome, "observacao": r.observacao} for r in obs_rows]

    nome_arquivo = f"inventario_consolidado_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    excel_bytes = _build_excel(itens, estoque_db, "Consolidado Geral", observacoes)

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )
