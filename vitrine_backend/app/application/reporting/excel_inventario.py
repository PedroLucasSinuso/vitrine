"""
app/application/reporting/excel_inventario.py

Geração de relatórios Excel para o módulo de inventário.
Extraído de app/api/routes/inventario.py para separar responsabilidades.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Estilos compartilhados ──────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", start_color="1E3A5F")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_ALT_FILL = PatternFill("solid", start_color="F4F7FB")
_TOTAL_FILL = PatternFill("solid", start_color="E8F0FE")
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_RED_FONT = Font(name="Arial", size=10, color="C0392B", bold=True)
_GREEN_FONT = Font(name="Arial", size=10, color="1A7A3C", bold=True)
_NORMAL_FONT = Font(name="Arial", size=10)


def _header_cell(ws, row: int, col: int, value: str):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _HEADER_FONT
    c.fill = _HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _BORDER
    return c


def _data_cell(ws, row: int, col: int, value, center: bool = False, alt: bool = False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _NORMAL_FONT
    c.border = _BORDER
    c.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
    )
    if alt:
        c.fill = _ALT_FILL
    return c


# ── Gerador principal ───────────────────────────────────────────────────────


def build_excel_inventario(
    itens_contados: list[dict],
    estoque_db: dict[str, float],  # codigo → estoque atual
    nome_relatorio: str,
    observacoes: list[dict] | None = None,
) -> bytes:
    """
    Gera o .xlsx em memória e retorna os bytes.

    itens_contados : lista de {codigo, nome, grupo, familia, quantidade}
    estoque_db     : mapa codigo → estoque atual do sistema
    observacoes    : lista opcional de {codigo, nome, observacao}
                     (se vazia, a aba "Observações" não é criada)
    """
    wb = Workbook()

    # ── Aba 1: Contagem ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Contagem"
    ws1.freeze_panes = "A2"

    cols1 = ["Código", "Produto", "Grupo", "Família", "Qtd. Contada"]
    for col, h in enumerate(cols1, 1):
        _header_cell(ws1, 1, col, h)
    ws1.row_dimensions[1].height = 22

    for ri, item in enumerate(itens_contados, 2):
        alt = ri % 2 == 0
        _data_cell(ws1, ri, 1, item["codigo"], center=True, alt=alt)
        _data_cell(ws1, ri, 2, item["nome"], alt=alt)
        _data_cell(ws1, ri, 3, item["grupo"], alt=alt)
        _data_cell(ws1, ri, 4, item["familia"], alt=alt)
        _data_cell(ws1, ri, 5, item["quantidade"], center=True, alt=alt)

    last1 = len(itens_contados) + 1
    total_row1 = last1 + 1
    for col in range(1, 6):
        c = ws1.cell(row=total_row1, column=col)
        c.border = _BORDER
        c.fill = _TOTAL_FILL
        c.font = Font(bold=True, name="Arial", size=10)
        if col == 4:
            c.value = "TOTAL"
            c.alignment = Alignment(horizontal="right", vertical="center")
        elif col == 5:
            c.value = f"=SUM(E2:E{last1})"
            c.alignment = Alignment(horizontal="center", vertical="center")

    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 34
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 16
    ws1.column_dimensions["E"].width = 14

    # ── Aba 2: Delta ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Delta (vs. Sistema)")
    ws2.freeze_panes = "A2"

    cols2 = [
        "Código", "Produto", "Grupo", "Família",
        "Estoque Sistema", "Qtd. Contada", "Delta", "Status",
    ]
    for col, h in enumerate(cols2, 1):
        _header_cell(ws2, 1, col, h)
    ws2.row_dimensions[1].height = 22

    for ri, item in enumerate(itens_contados, 2):
        alt = ri % 2 == 0
        cod = item["codigo"]
        contado = item["quantidade"]
        db_stock = estoque_db.get(cod)  # None se não encontrado no DB

        db_val: float | str = round(db_stock, 3) if db_stock is not None else "—"
        delta_formula = f"=F{ri}-E{ri}" if db_stock is not None else "—"

        if db_stock is None:
            status = "Sem cadastro"
        elif contado == 0:
            status = "Não contado"
        elif contado > db_stock:
            status = "Sobra"
        elif contado < db_stock:
            status = "Falta"
        else:
            status = "OK"

        _data_cell(ws2, ri, 1, cod, center=True, alt=alt)
        _data_cell(ws2, ri, 2, item["nome"], alt=alt)
        _data_cell(ws2, ri, 3, item["grupo"], alt=alt)
        _data_cell(ws2, ri, 4, item["familia"], alt=alt)
        _data_cell(ws2, ri, 5, db_val, center=True, alt=alt)
        _data_cell(ws2, ri, 6, contado, center=True, alt=alt)

        # Célula delta — fórmula colorida
        c_delta = ws2.cell(row=ri, column=7, value=delta_formula)
        c_delta.border = _BORDER
        c_delta.alignment = Alignment(horizontal="center", vertical="center")
        if alt:
            c_delta.fill = _ALT_FILL
        if db_stock is not None:
            delta = contado - db_stock
            if delta > 0:
                c_delta.font = _GREEN_FONT
            elif delta < 0:
                c_delta.font = _RED_FONT
            else:
                c_delta.font = Font(name="Arial", size=10, color="555555")
        else:
            c_delta.font = Font(name="Arial", size=10, color="AAAAAA")

        # Célula status — colorida
        c_status = ws2.cell(row=ri, column=8, value=status)
        c_status.border = _BORDER
        c_status.alignment = Alignment(horizontal="center", vertical="center")
        if alt:
            c_status.fill = _ALT_FILL
        if status == "OK":
            c_status.font = _GREEN_FONT
        elif status in ("Falta", "Não contado", "Sem cadastro"):
            c_status.font = _RED_FONT
        elif status == "Sobra":
            c_status.font = Font(name="Arial", size=10, color="D4800A", bold=True)
        else:
            c_status.font = _NORMAL_FONT

    # Totais da aba Delta
    last2 = len(itens_contados) + 1
    total_row2 = last2 + 1
    numeric_rows = [
        i for i, item in enumerate(itens_contados, 2)
        if estoque_db.get(item["codigo"]) is not None
    ]
    for col in range(1, 9):
        c = ws2.cell(row=total_row2, column=col)
        c.border = _BORDER
        c.fill = _TOTAL_FILL
        c.font = Font(bold=True, name="Arial", size=10)
        if col == 4:
            c.value = "TOTAL"
            c.alignment = Alignment(horizontal="right", vertical="center")
        elif col == 5:
            c.value = f"=SUM(E2:E{last2})"
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 6:
            c.value = f"=SUM(F2:F{last2})"
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 7:
            c.value = f"=SUM(G2:G{last2})"
            c.alignment = Alignment(horizontal="center", vertical="center")

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 34
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 17
    ws2.column_dimensions["F"].width = 15
    ws2.column_dimensions["G"].width = 10
    ws2.column_dimensions["H"].width = 15

    # ── Aba 3: Observações ───────────────────────────────────────────
    if observacoes:
        ws3 = wb.create_sheet("Observações")
        ws3.freeze_panes = "A2"

        cols3 = ["Código", "Produto", "Observação"]
        for col, h in enumerate(cols3, 1):
            _header_cell(ws3, 1, col, h)
        ws3.row_dimensions[1].height = 22

        for ri, obs in enumerate(observacoes, 2):
            alt = ri % 2 == 0
            _data_cell(ws3, ri, 1, obs.get("codigo", ""), center=True, alt=alt)
            _data_cell(ws3, ri, 2, obs.get("nome", ""), alt=alt)
            _data_cell(ws3, ri, 3, obs.get("observacao", ""), alt=alt)

        ws3.column_dimensions["A"].width = 18
        ws3.column_dimensions["B"].width = 34
        ws3.column_dimensions["C"].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
