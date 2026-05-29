"""
app/application/reporting/excel_margem_negativa.py

Geração de relatório Excel para produtos com margem negativa (pós-sync).
Segue o mesmo design de excel_inventario.py (cores, fontes, layout).
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Estilos compartilhados (mesmo design do inventário) ─────────────────────

_HEADER_FILL = PatternFill("solid", start_color="1E3A5F")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_ALT_FILL = PatternFill("solid", start_color="F4F7FB")
_TOTAL_FILL = PatternFill("solid", start_color="E8F0FE")
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_RED_FONT = Font(name="Arial", size=10, color="C0392B", bold=True)
_GREEN_FONT = Font(name="Arial", size=10, color="1A7A3C", bold=True)
_NORMAL_FONT = Font(name="Arial", size=10)


def _header_cell(ws, row: int, col: int, value: str):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _HEADER_FONT
    c.fill = _HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _BORDER
    return c


def _data_cell(ws, row: int, col: int, value, center: bool = False, alt: bool = False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _NORMAL_FONT
    c.border = _BORDER
    c.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
    )
    if alt:
        c.fill = _ALT_FILL
    return c


# ── Gerador principal ───────────────────────────────────────────────────────


def build_excel_margem_negativa(
    itens: list[dict],
    nome_relatorio: str = "Margem Negativa",
) -> bytes:
    """Gera .xlsx com os produtos de margem negativa.

    itens : lista de dicts com:
        - codigo         (str)
        - nome           (str)
        - grupo          (str)
        - preco_venda    (float)
        - preco_custo    (float)
        - margem         (float)  — percentual negativo
        - quantidade     (int)
        - receita        (float)  — receita dos últimos 30 dias (apenas produtos ativos em 90d)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Margem Negativa"
    ws.freeze_panes = "A2"

    # ── Cabeçalho com período dinâmico ───────────────────────────────────────
    from app.application.triggers_pos_sync import DIAS_ANALISE
    periodo_receita = f"Receita ({DIAS_ANALISE}d)"

    cols = [
        ("Código", 18),
        ("Produto", 38),
        ("Grupo", 16),
        ("Preço Venda", 14),
        ("Preço Custo", 14),
        ("Margem %", 11),
        ("Estoque", 10),
        (periodo_receita, 16),
        ("Impacto Mensal", 18),
    ]
    for col, (h, w) in enumerate(cols, 1):
        ws.column_dimensions[chr(64 + col)].width = w
        _header_cell(ws, 1, col, h)
    ws.row_dimensions[1].height = 22

    # ── Formato moeda pt-BR: R$ 1.234,56 ──────────────────────────────────
    BRL = '#,##0.00'

    for ri, item in enumerate(itens, 2):
        alt = ri % 2 == 0
        venda = round(float(item.get("preco_venda", 0)), 2)
        custo = round(float(item.get("preco_custo", 0)), 2)
        qtd = int(item.get("quantidade", 0))
        receita = round(float(item.get("receita", 0)), 2)
        impacto = round(qtd * (custo - venda), 2)

        _data_cell(ws, ri, 1, item.get("codigo", ""), center=True, alt=alt)
        _data_cell(ws, ri, 2, item.get("nome", ""), alt=alt)
        _data_cell(ws, ri, 3, item.get("grupo", ""), alt=alt)

        # Preço Venda — moeda
        c = ws.cell(row=ri, column=4, value=venda)
        c.number_format = BRL
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        if alt:
            c.fill = _ALT_FILL

        # Preço Custo — moeda
        c = ws.cell(row=ri, column=5, value=custo)
        c.number_format = BRL
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        if alt:
            c.fill = _ALT_FILL

        # Margem % — vermelho
        c = ws.cell(row=ri, column=6, value=round(float(item.get("margem", 0)), 2))
        c.number_format = '0.00"%"'
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = _RED_FONT
        if alt:
            c.fill = _ALT_FILL

        # Estoque — inteiro
        _data_cell(ws, ri, 7, qtd, center=True, alt=alt)

        # Receita — moeda
        c = ws.cell(row=ri, column=8, value=receita)
        c.number_format = BRL
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        if alt:
            c.fill = _ALT_FILL

        # Impacto — moeda vermelho
        c = ws.cell(row=ri, column=9, value=impacto)
        c.number_format = BRL
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = _RED_FONT
        if alt:
            c.fill = _ALT_FILL

    # ── Linha de total ─────────────────────────────────────────────────────
    last = len(itens) + 1
    total_row = last + 1
    for col in range(1, 10):
        c = ws.cell(row=total_row, column=col)
        c.border = _BORDER
        c.fill = _TOTAL_FILL
        c.font = Font(bold=True, name="Arial", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        if col == 1:
            c.value = "TOTAL"
            c.alignment = Alignment(horizontal="right", vertical="center")
        elif col == 4:
            c.value = f"=AVERAGE(D2:D{last})"
            c.number_format = BRL
        elif col == 7:
            c.value = f"=SUM(G2:G{last})"
        elif col == 8:
            c.value = f"=SUM(H2:H{last})"
            c.number_format = BRL
        elif col == 9:
            c.value = f"=SUM(I2:I{last})"
            c.number_format = BRL

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
