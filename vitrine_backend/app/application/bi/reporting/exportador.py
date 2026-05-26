"""
app/application/bi/reporting/exportador.py

Exportador de relatórios BI para Excel (.xlsx).
Versão estilizada com formatação automática de números,
largura de colunas adaptativa e linhas alternadas.
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter


# ── Paleta de cores ──────────────────────────────────────────────────────────
COR_HEADER = "059669"        # verde escuro
COR_HEADER_FONT = "FFFFFF"
COR_ALT = "F0FDF4"           # verde claro (alternado)
COR_BORDER = "D1D5DB"
COR_TOTAL_BG = "DCFCE7"      # fundo da linha de total

# ── Estilos compartilhados ──────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color=COR_HEADER_FONT, size=11)
_HEADER_FILL = PatternFill("solid", start_color=COR_HEADER, end_color=COR_HEADER)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BORDER = Border(
    left=Side(style="thin", color=COR_BORDER),
    right=Side(style="thin", color=COR_BORDER),
    top=Side(style="thin", color=COR_BORDER),
    bottom=Side(style="thin", color=COR_BORDER),
)
_ALT_FILL = PatternFill("solid", start_color=COR_ALT, end_color=COR_ALT)
_TOTAL_FILL = PatternFill("solid", start_color=COR_TOTAL_BG, end_color=COR_TOTAL_BG)
_TOTAL_FONT = Font(bold=True, size=11)

def _estilizar_planilha_bi(ws, cabecalhos: list[str], linhas: list[list]) -> None:
    """Aplica estilo completo a uma planilha BI."""
    # ── Cabeçalho ──
    for col_idx, cab in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col_idx, value=cab)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _BORDER

    # ── Dados ──
    for ri, linha in enumerate(linhas, 2):
        alt = ri % 2 == 0
        for col_idx, valor in enumerate(linha, 1):
            cell = ws.cell(row=ri, column=col_idx, value=valor)
            cell.border = _BORDER

            # Alinhamento
            if isinstance(valor, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Linhas alternadas
            if alt:
                cell.fill = _ALT_FILL

    # ── Congelar topo ──
    ws.freeze_panes = "A2"

    # ── Auto-filtro ──
    if linhas:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cabecalhos))}{len(linhas) + 1}"


def _adicionar_linha_total(ws, cabecalhos: list[str], linhas: list[list],
                            colunas_numericas: list[int]) -> None:
    """Adiciona uma linha de TOTAL ao final da planilha."""
    total_row = len(linhas) + 2  # +1 header, +1 data start
    for col_idx in range(1, len(cabecalhos) + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.border = _BORDER
        cell.fill = _TOTAL_FILL
        cell.font = _TOTAL_FONT

        if col_idx in colunas_numericas:
            col_letter = get_column_letter(col_idx)
            cell.value = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif col_idx == len(cabecalhos):  # última coluna não-numérica
            cell.value = "TOTAL"
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")


class ExportadorExcel:
    """Exportador de relatórios BI para Excel (.xlsx)."""

    def exportar(self, dados: dict[str, list[dict]]) -> bytes:
        """Gera .xlsx a partir de um dicionário {nome_aba: [dict, ...]}."""
        wb = Workbook()
        wb.remove(wb.active)

        for nome_aba, linhas in dados.items():
            if not linhas:
                continue
            ws = wb.create_sheet(title=nome_aba[:31])
            cabecalhos = list(linhas[0].keys())
            dados_linhas = [[linha.get(h, "") for h in cabecalhos] for linha in linhas]

            _estilizar_planilha_bi(ws, cabecalhos, dados_linhas)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def exportar_ranking(self, linhas: list[dict]) -> bytes:
        """Exporta ranking com destaque no top 3 e coluna de posição."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Ranking"

        cabecalhos = list(linhas[0].keys()) if linhas else []
        if not cabecalhos:
            wb.remove(ws)
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf.read()

        dados_linhas = [[linha.get(h, "") for h in cabecalhos] for linha in linhas]
        _estilizar_planilha_bi(ws, cabecalhos, dados_linhas)

        # ── Destaque top 3: ouro, prata, bronze ──
        medal_fills = {
            1: PatternFill("solid", start_color="FEE2A0", end_color="FEE2A0"),  # ouro
            2: PatternFill("solid", start_color="E5E7EB", end_color="E5E7EB"),  # prata
            3: PatternFill("solid", start_color="FDE3C8", end_color="FDE3C8"),  # bronze
        }
        for ri in range(2, min(5, len(linhas) + 2)):
            fill = medal_fills.get(ri - 1)
            if fill:
                for col_idx in range(1, len(cabecalhos) + 1):
                    ws.cell(row=ri, column=col_idx).fill = fill

        # Identificar coluna de posição/rank
        rank_col = None
        for i, cab in enumerate(cabecalhos, 1):
            if cab.lower() in ("posicao", "rank", "pos", "#"):
                rank_col = i
                break
        if rank_col is None:
            # Inserir coluna de posição à esquerda
            rank_col = 1
            ws.insert_cols(1)
            cabecalhos.insert(0, "Posição")
            for ri in range(2, len(linhas) + 2):
                cell = ws.cell(row=ri, column=1, value=ri - 1)
                cell.border = _BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if ri <= 4:
                    fill = medal_fills.get(ri - 1)
                    if fill:
                        cell.fill = fill
            # Atualizar header da coluna 1
            hcell = ws.cell(row=1, column=1, value="Posição")
            hcell.font = _HEADER_FONT
            hcell.fill = _HEADER_FILL
            hcell.alignment = _HEADER_ALIGN
            hcell.border = _BORDER
            # Atualizar auto-filtro
            ws.auto_filter.ref = f"A1:{get_column_letter(len(cabecalhos))}{len(linhas) + 1}"

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def exportar_diario(self, linhas: list[dict]) -> bytes:
        """Exporta série diária com linha de totais."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Série Diária"

        cabecalhos = list(linhas[0].keys()) if linhas else []
        if not cabecalhos:
            wb.remove(ws)
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf.read()

        dados_linhas = [[linha.get(h, "") for h in cabecalhos] for linha in linhas]
        _estilizar_planilha_bi(ws, cabecalhos, dados_linhas)

        # Adicionar linha de total em todas as colunas exceto a primeira (data)
        colunas_num = list(range(2, len(cabecalhos) + 1)) if len(cabecalhos) > 1 else []
        if colunas_num:
            _adicionar_linha_total(ws, cabecalhos, dados_linhas, colunas_num)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()


# ── Função auxiliar para anexo de relatório semanal ──────────────────────────


def build_bi_anexo(
    kpis_semana: dict,
    kpis_mes: dict,
    ranking: list[dict],
) -> bytes:
    """
    Gera .xlsx com 3 abas (Semana, Mês, Ranking) para anexar ao email semanal.
    Aceita dicionários (model_dump()) em vez de objetos para evitar acoplamento.
    """
    wb = Workbook()

    # ── Aba 1: Semana ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Semana"
    cab_week = list(kpis_semana.keys())
    _estilizar_planilha_bi(ws1, cab_week, [list(kpis_semana.values())])

    # ── Aba 2: Mês ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Mês")
    cab_month = list(kpis_mes.keys())
    _estilizar_planilha_bi(ws2, cab_month, [list(kpis_mes.values())])

    # ── Aba 3: Ranking ──────────────────────────────────────────────
    if ranking:
        ws3 = wb.create_sheet("Ranking")
        cab_rank = list(ranking[0].keys())
        dados_rank = [list(r.values()) for r in ranking]
        _estilizar_planilha_bi(ws3, cab_rank, dados_rank)

        # Destaque top 3
        medal_fills = {
            0: PatternFill("solid", start_color="FEE2A0", end_color="FEE2A0"),
            1: PatternFill("solid", start_color="E5E7EB", end_color="E5E7EB"),
            2: PatternFill("solid", start_color="FDE3C8", end_color="FDE3C8"),
        }
        for i in range(min(3, len(ranking))):
            fill = medal_fills.get(i)
            if fill:
                for col in range(1, len(cab_rank) + 1):
                    ws3.cell(row=i + 2, column=col).fill = fill

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
