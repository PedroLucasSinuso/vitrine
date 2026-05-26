"""
Testes para o ExportadorExcel estilizado + build_bi_anexo.
"""

import pytest
from openpyxl import load_workbook
from io import BytesIO
from app.application.bi.reporting.exportador import ExportadorExcel, build_bi_anexo


class TestExportadorExcel:
    """Testes para o exportador base (estilizado)."""

    def test_exportar_multiplas_abas(self):
        """Deve criar múltiplas abas."""
        dados = {
            "Aba1": [{"nome": "João", "valor": 100.50}],
            "Aba2": [{"codigo": "ABC", "qtd": 5}],
        }
        exportador = ExportadorExcel()
        buf = exportador.exportar(dados)
        wb = load_workbook(BytesIO(buf))
        assert "Aba1" in wb.sheetnames
        assert "Aba2" in wb.sheetnames
        assert len(wb.sheetnames) == 2

    def test_exportar_aba_vazia_ignorada(self):
        """Abas sem dados não devem ser criadas."""
        dados = {"Vazia": [], "Cheia": [{"x": 1}]}
        exportador = ExportadorExcel()
        buf = exportador.exportar(dados)
        wb = load_workbook(BytesIO(buf))
        assert "Vazia" not in wb.sheetnames
        assert "Cheia" in wb.sheetnames

    def test_valores_preservados(self):
        """Valores devem ser preservados sem formatação numérica adicional."""
        dados = {
            "KPIs": [
                {"faturamento_bruto": 1234.56, "qtd_tickets": 42},
            ],
        }
        exportador = ExportadorExcel()
        buf = exportador.exportar(dados)
        wb = load_workbook(BytesIO(buf))
        ws = wb.active
        # Valores devem estar presentes e sem formatação numérica automática
        assert ws.cell(row=2, column=1).value == 1234.56
        assert ws.cell(row=2, column=2).value == 42
        # Formato deve ser 'General' (padrão do openpyxl) — não aplicamos formatos
        assert ws.cell(row=2, column=1).number_format == 'General'

    def test_cabecalho_congelado(self):
        """Primeira linha deve estar congelada (freeze_panes = A2)."""
        dados = {"Teste": [{"a": 1}]}
        exportador = ExportadorExcel()
        buf = exportador.exportar(dados)
        wb = load_workbook(BytesIO(buf))
        ws = wb.active
        assert ws.freeze_panes == "A2"

    def test_auto_filtro_presente(self):
        """Auto-filtro deve estar ativo."""
        dados = {"Teste": [{"col1": "val1", "col2": 123}]}
        exportador = ExportadorExcel()
        buf = exportador.exportar(dados)
        wb = load_workbook(BytesIO(buf))
        ws = wb.active
        assert ws.auto_filter.ref is not None


class TestExportadorRanking:
    """Testes para exportar_ranking."""

    def test_ranking_adiciona_coluna_posicao(self):
        """Se não houver coluna de posição, deve inseri-la."""
        linhas = [
            {"produto": "Item A", "valor": 1000.0},
            {"produto": "Item B", "valor": 800.0},
            {"produto": "Item C", "valor": 600.0},
        ]
        exportador = ExportadorExcel()
        buf = exportador.exportar_ranking(linhas)
        wb = load_workbook(BytesIO(buf))
        ws = wb.active
        # Primeira coluna deve ser "Posição"
        assert ws.cell(row=1, column=1).value == "Posição"
        # Primeiro item deve ser posição 1
        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=3, column=1).value == 2

    def test_ranking_top3_colorido(self):
        """Top 3 deve ter preenchimento especial."""
        linhas = [
            {"produto": "A", "valor": 100},
            {"produto": "B", "valor": 80},
            {"produto": "C", "valor": 60},
            {"produto": "D", "valor": 40},
        ]
        exportador = ExportadorExcel()
        buf = exportador.exportar_ranking(linhas)
        wb = load_workbook(BytesIO(buf))
        ws = wb.active
        # Top 3 (linhas 2-4) devem ter fill diferente da linha 5
        fill_top1 = ws.cell(row=2, column=1).fill
        fill_top3 = ws.cell(row=4, column=1).fill
        fill_pos4 = ws.cell(row=5, column=1).fill
        assert fill_top1.start_color.rgb != fill_pos4.start_color.rgb
        assert fill_top3.start_color.rgb != fill_pos4.start_color.rgb


class TestExportadorDiario:
    """Testes para exportar_diario."""

    def test_diario_linha_total(self):
        """Deve adicionar linha de TOTAL para colunas numéricas."""
        linhas = [
            {"data": "2026-01-01", "receita": 1000.0, "qtd": 10},
            {"data": "2026-01-02", "receita": 1500.0, "qtd": 15},
        ]
        exportador = ExportadorExcel()
        buf = exportador.exportar_diario(linhas)
        wb = load_workbook(BytesIO(buf))
        ws = wb.active
        # Última linha deve ser TOTAL
        total_row = len(linhas) + 2  # +1 header, +1 data
        # Coluna 2 (receita) deve conter fórmula SUM
        assert ws.cell(row=total_row, column=2).value is not None
        assert "SUM" in str(ws.cell(row=total_row, column=2).value)


class TestBuildBiAnexo:
    """Testes para build_bi_anexo (anexo do email semanal)."""

    def test_gerar_anexo_completo(self):
        """Deve gerar .xlsx com 3 abas."""
        kpis_semana = {"faturamento": 5000.0, "tickets": 50}
        kpis_mes = {"faturamento": 20000.0, "tickets": 200}
        ranking = [
            {"produto": "Top1", "valor": 5000},
            {"produto": "Top2", "valor": 3000},
        ]
        buf = build_bi_anexo(kpis_semana, kpis_mes, ranking)
        wb = load_workbook(BytesIO(buf))
        assert "Semana" in wb.sheetnames
        assert "Mês" in wb.sheetnames
        assert "Ranking" in wb.sheetnames

    def test_anexo_sem_ranking(self):
        """Se ranking vazio, aba Ranking não deve ser criada."""
        kpis_semana = {"faturamento": 5000.0, "tickets": 50}
        kpis_mes = {"faturamento": 20000.0, "tickets": 200}
        buf = build_bi_anexo(kpis_semana, kpis_mes, [])
        wb = load_workbook(BytesIO(buf))
        assert "Semana" in wb.sheetnames
        assert "Mês" in wb.sheetnames
        assert "Ranking" not in wb.sheetnames
